import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#def process_workbook(filename):
wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
#cell = sheet["a1"]  ----How to accss a cell
#cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    correct_price = cell.value * 0.9
    correct_price_cell = sheet.cell(row, 4)
    correct_price_cell.value = correct_price

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "f2")

wb.save("transactions2.xlsx")
wb.save("transactions3.xlsx")
